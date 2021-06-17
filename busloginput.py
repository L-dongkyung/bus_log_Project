import sys, os, re
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.QtCore import QDate
from openpyxl import Workbook, load_workbook
from datetime import datetime

form_class = uic.loadUiType("./record.ui")[0]


class Windowclass(QMainWindow, form_class): # 위젯 클래스

    class exel():
        def __init__(self, xl_name):
            file_path = xl_name
            now_path = os.path.dirname(os.path.realpath(__file__))
            if os.path.exists(f"{now_path}/{file_path}.xlsx"):  # 기존 엑셀 로드
                wb = load_workbook(f"{now_path}/{file_path}.xlsx")
                self.wb = wb
            else:
                wb = load_workbook(f"{now_path}/backup.xlsx")  # 새 엑셀 생성
                self.wb = wb

        def select_sheet(self, region):
            ws = self.wb[region]
            self.ws = ws

        def write_xl(self, date, region, start_hour, start_min, end_hour, end_min, meter, people, etc):
            if self.ws.max_row != 5 and self.ws[f'A{self.ws.max_row}'].value != None:
                last_no = self.ws[f'A{self.ws.max_row}'].value
            else:
                last_no = 0
            start_time = datetime.strptime(f"{start_hour}:{start_min}", "%H:%M")
            end_time = datetime.strptime(f"{end_hour}:{end_min}", "%H:%M")
            time = str(end_time-start_time)[:-3]
            if last_no != 0:
                pre_meter = self.ws[f'H{self.ws.max_row}'].value
                ride = int(meter) - int(pre_meter)
            else:
                ride = 0

            self.ws.append([last_no+1, f'{region}', f'{date}', f"{start_hour}:{start_min:0>2}", "~",
                            f"{end_hour}:{end_min:0>2}", f'{time}', int(meter), ride, int(people), etc])

        def save_xl(self):
            self.wb.save('기업명 통근버스 및 민원인버스 운행일지(통합).xlsx')
            self.wb.close()

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.setFixedSize(618, 590) # 창 크기 고정
        self.pushButton_save.clicked.connect(self.saveEvent)  # 저장 버튼 클릭시 알림
        self.pushButton_cancle.clicked.connect(self.close)  # 취소 버튼 클릭시 종료
        self.dateEdit.setDate(QDate.currentDate())  # 현재 날짜 출력
        self.xl = self.exel('기업명 통근버스 및 민원인버스 운행일지(통합)')

    # 저장버튼 클릭시 알림 출력 후 종료
    def saveEvent(self):
        date = self.dateEdit.date().toString("yyyy년 MM월 dd일 dddd")  # 선택된 날짜(디폴트 = 현재날짜)
        region = self.comboBox_region.currentText()  # 콤보박스의 현재 지역
        start_hour = self.lineEdit_start_hour.text()
        start_min = self.lineEdit_start_min.text()
        end_hour = self.lineEdit_end_hour.text()
        end_min = self.lineEdit_end_min.text()
        meter = self.lineEdit_meter.text()
        people = self.lineEdit_people.text()
        etc = self.plainTextEdit_ETC.toPlainText()
        self.xl.select_sheet(region)
        try:
            self.xl.write_xl(date, region,start_hour, start_min, end_hour, end_min, meter, people, etc)
        except:
            except_msg = QMessageBox()
            except_msg.setText("저장에 실패했습니다.\n입력내용을 확인하세요")
            except_msg.exec_()
            return
        self.xl.save_xl()
        msgbox = QMessageBox()
        msgbox.setText("저장이 완료되었습니다.\n잘못 입력하였을 경우 담당자에게 연락하세요.\n"
                       "다시 입력하시고 비고에 다음행은 잘못 입력했다고 표시하세요.")
        msgbox.addButton("확인", QMessageBox.YesRole)  # 0
        msgbox.exec_()
        self.close()




app = QApplication(sys.argv)
mainwindow = Windowclass()
mainwindow.show()
app.exec()
